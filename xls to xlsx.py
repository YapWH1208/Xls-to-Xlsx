import xlrd
from openpyxl.workbook import Workbook


def open_xls_as_xlsx(xls_path, xlsx_path):
    # first open file using xlrd
    book = xlrd.open_workbook(xls_path)
    index = 0
    nrows, ncols = 0, 0
    sheet = book.sheet_by_index(0)
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book_new = Workbook()
    sheet_new = book_new.create_sheet("sheet1", 0)

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet_new.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    book_new.save(xlsx_path)

open_xls_as_xlsx("./123.xls", "./123.xlsx")

